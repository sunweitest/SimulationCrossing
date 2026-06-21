"""管理后台 API

提供用户查询、追加次数、包月不限次数等功能。
使用 Redis 存储额外配额和包月标记。

新增：小说/时间节点/预设角色 CRUD 管理。
新增：角色批量导入（xlsx 模板下载/上传）。
"""
import logging
import io
from fastapi import APIRouter, Depends, HTTPException, Header, status, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from sqlalchemy.orm import selectinload
from app.core.database import get_db
from app.core.config import settings
from app.models.models import User, GameSession, PresetCharacter, CharacterTimeline, NovelConfig, TimelineConfig
from app.schemas.schemas import (
    NovelCreate, NovelUpdate, NovelResponse,
    TimelineCreate, TimelineUpdate, TimelineResponse,
    PresetCharacterAdminCreate, PresetCharacterAdminUpdate, PresetCharacterAdminResponse,
    TimelineEntrySchema,
)
from datetime import date, datetime
from typing import Optional
import redis.asyncio as redis

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

logger = logging.getLogger("admin_api")

router = APIRouter(prefix="/api/admin", tags=["管理后台"])

_redis: redis.Redis = None


def _get_redis():
    global _redis
    if _redis is None:
        _redis = redis.Redis(
            host=settings.REDIS_HOST,
            port=settings.REDIS_PORT,
            db=settings.REDIS_DB,
            decode_responses=True,
        )
    return _redis


async def _verify_admin(x_admin_key: str = Header(...)):
    """验证管理后台密钥"""
    if x_admin_key != settings.ADMIN_KEY:
        raise HTTPException(status_code=403, detail="无效的管理密钥")
    return True


def _extra_quota_key(user_id: int) -> str:
    return f"rate_limit:user:{user_id}:extra"


def _unlimited_key(user_id: int) -> str:
    """包月不限次数标记，30 天过期"""
    return f"rate_limit:user:{user_id}:unlimited"


# ── 用户查询 ──────────────────────────────────────────────

@router.get("/lookup")
async def lookup_user(
    q: str,
    db: AsyncSession = Depends(get_db),
    _: bool = Depends(_verify_admin),
):
    """根据邮箱或手机号查找用户"""
    result = await db.execute(
        select(User).where((User.email == q) | (User.phone == q))
    )
    user = result.scalar_one_or_none()

    if not user:
        raise HTTPException(status_code=404, detail="未找到该用户")

    r = _get_redis()

    # 查询当前额外配额和包月状态
    extra = await r.get(_extra_quota_key(user.id)) or "0"
    unlimited_key = _unlimited_key(user.id)
    unlimited_ttl = await r.ttl(unlimited_key)

    return {
        "id": user.id,
        "email": user.email,
        "phone": user.phone,
        "is_active": user.is_active,
        "created_at": user.created_at.isoformat(),
        "extra_quota": int(extra),
        "is_unlimited": unlimited_ttl > 0,
        "unlimited_days_left": max(0, unlimited_ttl // 86400) if unlimited_ttl > 0 else 0,
    }


# ── 追加次数 ──────────────────────────────────────────────

@router.post("/add-quota")
async def add_quota(
    user_id: int,
    amount: int,
    db: AsyncSession = Depends(get_db),
    _: bool = Depends(_verify_admin),
):
    """给指定用户追加额外次数"""
    if amount <= 0:
        raise HTTPException(status_code=400, detail="次数必须大于0")

    # 确认用户存在
    user = await db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")

    r = _get_redis()
    key = _extra_quota_key(user_id)
    new_total = await r.incrby(key, amount)
    # 额外配额 30 天过期
    await r.expire(key, 86400 * 30)
    # 清除当天的计数缓存，让用户立即生效
    today_key = f"rate_limit:user:{user_id}:{date.today()}"
    await r.delete(today_key)

    logger.info("ADMIN: add_quota user=%s amount=%d total_extra=%d", user_id, amount, new_total)

    return {
        "message": f"已为用户 {user.email or user.phone} 追加 {amount} 次",
        "extra_quota": new_total,
    }


# ── 包月不限次数 ──────────────────────────────────────────

@router.post("/monthly-unlimited")
async def set_monthly_unlimited(
    user_id: int,
    db: AsyncSession = Depends(get_db),
    _: bool = Depends(_verify_admin),
):
    """给指定用户开通本月不限次数"""
    user = await db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")

    r = _get_redis()
    key = _unlimited_key(user_id)

    # 30 天包月
    ttl = 30 * 86400

    await r.set(key, "1", ex=ttl)
    # 清除当天的计数缓存
    today_key = f"rate_limit:user:{user_id}:{date.today()}"
    await r.delete(today_key)

    logger.info("ADMIN: monthly_unlimited user=%s days=30", user_id)

    return {
        "message": f"已为用户 {user.email or user.phone} 开通30天不限次数",
        "days_left": 30,
    }


@router.delete("/monthly-unlimited")
async def cancel_monthly_unlimited(
    user_id: int,
    db: AsyncSession = Depends(get_db),
    _: bool = Depends(_verify_admin),
):
    """取消用户的本月不限次数"""
    user = await db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")

    r = _get_redis()
    key = _unlimited_key(user_id)
    await r.delete(key)

    logger.info("ADMIN: cancel_unlimited user=%s", user_id)

    return {"message": f"已取消用户 {user.email or user.phone} 的本月不限次数"}


# ═══════════════════════════════════════════════════════════════
# 小说/历史背景 CRUD
# ═══════════════════════════════════════════════════════════════

@router.get("/novels", response_model=list[NovelResponse])
async def list_novels(db: AsyncSession = Depends(get_db), _: bool = Depends(_verify_admin)):
    """列出所有小说"""
    result = await db.execute(
        select(NovelConfig)
        .options(selectinload(NovelConfig.timelines))
        .order_by(NovelConfig.sort_order, NovelConfig.id)
    )
    novels = result.scalars().all()
    return [
        NovelResponse(
            id=n.id,
            name=n.name,
            description=n.description,
            sort_order=n.sort_order,
            timeline_count=len(n.timelines) if n.timelines else 0,
            character_count=0,
            created_at=n.created_at,
        )
        for n in novels
    ]


@router.post("/novels", response_model=NovelResponse, status_code=201)
async def create_novel(
    data: NovelCreate,
    db: AsyncSession = Depends(get_db),
    _: bool = Depends(_verify_admin),
):
    """创建小说"""
    existing = await db.execute(select(NovelConfig).where(NovelConfig.name == data.name))
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=409, detail="该小说名称已存在")
    novel = NovelConfig(name=data.name, description=data.description, sort_order=data.sort_order)
    db.add(novel)
    await db.commit()
    await db.refresh(novel)
    return NovelResponse(
        id=novel.id, name=novel.name, description=novel.description,
        sort_order=novel.sort_order, timeline_count=0, character_count=0, created_at=novel.created_at,
    )


@router.get("/novels/{novel_id}", response_model=NovelResponse)
async def get_novel(
    novel_id: int,
    db: AsyncSession = Depends(get_db),
    _: bool = Depends(_verify_admin),
):
    """获取小说详情（含时间节点列表）"""
    result = await db.execute(
        select(NovelConfig).options(selectinload(NovelConfig.timelines)).where(NovelConfig.id == novel_id)
    )
    novel = result.scalar_one_or_none()
    if not novel:
        raise HTTPException(status_code=404, detail="小说不存在")
    return NovelResponse(
        id=novel.id, name=novel.name, description=novel.description,
        sort_order=novel.sort_order,
        timeline_count=len(novel.timelines) if novel.timelines else 0,
        character_count=0,
        created_at=novel.created_at,
    )


@router.put("/novels/{novel_id}", response_model=NovelResponse)
async def update_novel(
    novel_id: int,
    data: NovelUpdate,
    db: AsyncSession = Depends(get_db),
    _: bool = Depends(_verify_admin),
):
    """更新小说"""
    result = await db.execute(
        select(NovelConfig).options(selectinload(NovelConfig.timelines)).where(NovelConfig.id == novel_id)
    )
    novel = result.scalar_one_or_none()
    if not novel:
        raise HTTPException(status_code=404, detail="小说不存在")
    if data.name is not None:
        novel.name = data.name
    if data.description is not None:
        novel.description = data.description
    if data.sort_order is not None:
        novel.sort_order = data.sort_order
    await db.commit()
    await db.refresh(novel)
    return NovelResponse(
        id=novel.id, name=novel.name, description=novel.description,
        sort_order=novel.sort_order,
        timeline_count=len(novel.timelines) if novel.timelines else 0,
        character_count=0,
        created_at=novel.created_at,
    )


@router.delete("/novels/{novel_id}")
async def delete_novel(
    novel_id: int,
    db: AsyncSession = Depends(get_db),
    _: bool = Depends(_verify_admin),
):
    """删除小说（级联删除关联的时间节点；如有角色引用则拒绝）"""
    result = await db.execute(select(NovelConfig).where(NovelConfig.id == novel_id))
    novel = result.scalar_one_or_none()
    if not novel:
        raise HTTPException(status_code=404, detail="小说不存在")
    # 检查是否有角色引用此小说名称
    char_count = await db.execute(
        select(func.count(PresetCharacter.id)).where(PresetCharacter.novel == novel.name)
    )
    if char_count.scalar() > 0:
        raise HTTPException(status_code=409, detail=f"该小说下有 {char_count.scalar()} 个角色，请先删除角色")
    # 检查是否有游戏会话引用
    session_count = await db.execute(
        select(func.count(GameSession.id)).where(GameSession.novel == novel.name)
    )
    if session_count.scalar() > 0:
        raise HTTPException(status_code=409, detail=f"存在 {session_count.scalar()} 个游戏会话引用该小说，无法删除")
    await db.delete(novel)
    await db.commit()
    return {"message": "已删除"}


# ═══════════════════════════════════════════════════════════════
# 时间节点 CRUD
# ═══════════════════════════════════════════════════════════════

@router.get("/timelines", response_model=list[TimelineResponse])
async def list_timelines(
    novel_id: int = Query(...),
    db: AsyncSession = Depends(get_db),
    _: bool = Depends(_verify_admin),
):
    """列出某小说的所有时间节点"""
    result = await db.execute(
        select(TimelineConfig)
        .where(TimelineConfig.novel_id == novel_id)
        .order_by(TimelineConfig.sort_order, TimelineConfig.id)
    )
    timelines = result.scalars().all()
    return [
        TimelineResponse(
            id=t.id, novel_id=t.novel_id, novel_name="",
            name=t.name, description=t.description, sort_order=t.sort_order, created_at=t.created_at,
        )
        for t in timelines
    ]


@router.post("/timelines", response_model=TimelineResponse, status_code=201)
async def create_timeline(
    data: TimelineCreate,
    db: AsyncSession = Depends(get_db),
    _: bool = Depends(_verify_admin),
):
    """创建时间节点"""
    # 验证小说存在
    novel = await db.get(NovelConfig, data.novel_id)
    if not novel:
        raise HTTPException(status_code=404, detail="小说不存在")
    existing = await db.execute(
        select(TimelineConfig).where(
            TimelineConfig.novel_id == data.novel_id,
            TimelineConfig.name == data.name,
        )
    )
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=409, detail="该时间节点已存在")
    timeline = TimelineConfig(
        novel_id=data.novel_id, name=data.name,
        description=data.description, sort_order=data.sort_order,
    )
    db.add(timeline)
    await db.commit()
    await db.refresh(timeline)
    return TimelineResponse(
        id=timeline.id, novel_id=timeline.novel_id, novel_name=novel.name,
        name=timeline.name, description=timeline.description,
        sort_order=timeline.sort_order, created_at=timeline.created_at,
    )


@router.put("/timelines/{timeline_id}", response_model=TimelineResponse)
async def update_timeline(
    timeline_id: int,
    data: TimelineUpdate,
    db: AsyncSession = Depends(get_db),
    _: bool = Depends(_verify_admin),
):
    """更新时间节点"""
    timeline = await db.get(TimelineConfig, timeline_id)
    if not timeline:
        raise HTTPException(status_code=404, detail="时间节点不存在")
    if data.name is not None:
        timeline.name = data.name
    if data.description is not None:
        timeline.description = data.description
    if data.sort_order is not None:
        timeline.sort_order = data.sort_order
    await db.commit()
    await db.refresh(timeline)
    novel = await db.get(NovelConfig, timeline.novel_id)
    return TimelineResponse(
        id=timeline.id, novel_id=timeline.novel_id, novel_name=novel.name if novel else "",
        name=timeline.name, description=timeline.description,
        sort_order=timeline.sort_order, created_at=timeline.created_at,
    )


@router.delete("/timelines/{timeline_id}")
async def delete_timeline(
    timeline_id: int,
    db: AsyncSession = Depends(get_db),
    _: bool = Depends(_verify_admin),
):
    """删除时间节点（如有角色引用则拒绝）"""
    timeline = await db.get(TimelineConfig, timeline_id)
    if not timeline:
        raise HTTPException(status_code=404, detail="时间节点不存在")
    # 检查角色时间线引用
    ct_count = await db.execute(
        select(func.count(CharacterTimeline.id)).where(CharacterTimeline.timeline == timeline.name)
    )
    if ct_count.scalar() > 0:
        raise HTTPException(status_code=409, detail=f"有 {ct_count.scalar()} 个角色引用该时间节点，请先删除角色")
    await db.delete(timeline)
    await db.commit()
    return {"message": "已删除"}


# ═══════════════════════════════════════════════════════════════
# 预设角色 CRUD
# ═══════════════════════════════════════════════════════════════

@router.get("/characters", response_model=list[PresetCharacterAdminResponse])
async def list_admin_characters(
    novel: Optional[str] = None,
    timeline: Optional[str] = None,
    search: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    _: bool = Depends(_verify_admin),
):
    """列出预设角色（管理后台用，含完整时间线条目）"""
    stmt = select(PresetCharacter).options(selectinload(PresetCharacter.timelines))
    if novel:
        stmt = stmt.where(PresetCharacter.novel == novel)
    if search:
        stmt = stmt.where(PresetCharacter.name.ilike(f"%{search}%"))
    stmt = stmt.order_by(PresetCharacter.novel, PresetCharacter.id)
    result = await db.execute(stmt)
    characters = result.scalars().all()

    def filter_timelines(char):
        if not timeline:
            return char.timelines
        return [t for t in char.timelines if t.timeline == timeline]

    return [
        PresetCharacterAdminResponse(
            id=c.id, novel=c.novel, name=c.name,
            gender=c.gender, age=c.age, rank=c.rank,
            background=c.background, starting_points=c.starting_points,
            timelines=[
                TimelineEntrySchema(timeline=t.timeline, background=t.background, initial_scene=t.initial_scene)
                for t in filter_timelines(c)
            ],
        )
        for c in characters
    ]


@router.post("/characters", response_model=PresetCharacterAdminResponse, status_code=201)
async def create_admin_character(
    data: PresetCharacterAdminCreate,
    db: AsyncSession = Depends(get_db),
    _: bool = Depends(_verify_admin),
):
    """创建预设角色（含时间线条目）"""
    char = PresetCharacter(
        novel=data.novel, name=data.name,
        gender=data.gender, age=data.age, rank=data.rank,
        background=data.background, starting_points=data.starting_points,
    )
    for tl in data.timelines:
        char.timelines.append(CharacterTimeline(
            timeline=tl.timeline, background=tl.background, initial_scene=tl.initial_scene,
        ))
    db.add(char)
    await db.commit()
    await db.refresh(char)
    return PresetCharacterAdminResponse(
        id=char.id, novel=char.novel, name=char.name,
        gender=char.gender, age=char.age, rank=char.rank,
        background=char.background, starting_points=char.starting_points,
        timelines=[
            TimelineEntrySchema(timeline=t.timeline, background=t.background, initial_scene=t.initial_scene)
            for t in char.timelines
        ],
    )


@router.get("/characters/{char_id}", response_model=PresetCharacterAdminResponse)
async def get_admin_character(
    char_id: int,
    db: AsyncSession = Depends(get_db),
    _: bool = Depends(_verify_admin),
):
    """获取单个角色详情"""
    result = await db.execute(
        select(PresetCharacter).options(selectinload(PresetCharacter.timelines)).where(PresetCharacter.id == char_id)
    )
    char = result.scalar_one_or_none()
    if not char:
        raise HTTPException(status_code=404, detail="角色不存在")
    return PresetCharacterAdminResponse(
        id=char.id, novel=char.novel, name=char.name,
        gender=char.gender, age=char.age, rank=char.rank,
        background=char.background, starting_points=char.starting_points,
        timelines=[
            TimelineEntrySchema(timeline=t.timeline, background=t.background, initial_scene=t.initial_scene)
            for t in char.timelines
        ],
    )


@router.put("/characters/{char_id}", response_model=PresetCharacterAdminResponse)
async def update_admin_character(
    char_id: int,
    data: PresetCharacterAdminUpdate,
    db: AsyncSession = Depends(get_db),
    _: bool = Depends(_verify_admin),
):
    """更新角色（含时间线条目替换）"""
    result = await db.execute(
        select(PresetCharacter).options(selectinload(PresetCharacter.timelines)).where(PresetCharacter.id == char_id)
    )
    char = result.scalar_one_or_none()
    if not char:
        raise HTTPException(status_code=404, detail="角色不存在")

    # 更新基本字段
    for field in ("novel", "name", "gender", "age", "rank", "background", "starting_points"):
        val = getattr(data, field)
        if val is not None:
            setattr(char, field, val)

    # 替换时间线条目
    if data.timelines is not None:
        for old_tl in list(char.timelines):
            await db.delete(old_tl)
        for tl in data.timelines:
            new_tl = CharacterTimeline(
                character_id=char.id, timeline=tl.timeline,
                background=tl.background, initial_scene=tl.initial_scene,
            )
            db.add(new_tl)

    await db.commit()
    await db.refresh(char)
    return PresetCharacterAdminResponse(
        id=char.id, novel=char.novel, name=char.name,
        gender=char.gender, age=char.age, rank=char.rank,
        background=char.background, starting_points=char.starting_points,
        timelines=[
            TimelineEntrySchema(timeline=t.timeline, background=t.background, initial_scene=t.initial_scene)
            for t in char.timelines
        ],
    )


@router.delete("/characters/{char_id}")
async def delete_admin_character(
    char_id: int,
    db: AsyncSession = Depends(get_db),
    _: bool = Depends(_verify_admin),
):
    """删除角色"""
    char = await db.get(PresetCharacter, char_id)
    if not char:
        raise HTTPException(status_code=404, detail="角色不存在")
    await db.delete(char)
    await db.commit()
    return {"message": "已删除"}


# ═══════════════════════════════════════════════════════════════
# 角色批量导入（xlsx）— 必须在 /characters/{char_id} 之前注册
# ═══════════════════════════════════════════════════════════════

_XLSX_HEADERS = [
    "小说", "姓名", "性别", "年龄", "身份",
    "角色背景", "初始积分",
    "时间节点", "时间节点背景", "初始场景描述",
]

_XLSX_SAMPLE = [
    ["三国演义", "诸葛亮", "男性", 28, "军师",
     "字孔明，号卧龙，三国时期蜀汉丞相", 80,
     "赤壁之战", "与周瑜联手，借东风火烧曹营", "江边祭坛上，你手持羽扇，东南风渐起..."],
    ["三国演义", "诸葛亮", "男性", 28, "军师",
     "字孔明，号卧龙，三国时期蜀汉丞相", 80,
     "北伐中原", "六出祁山，鞠躬尽瘁", "祁山大营中，你看着地图上的五丈原..."],
]


@router.get("/character-template")
async def download_character_template(_: bool = Depends(_verify_admin)):
    """下载角色导入 xlsx 模板"""
    if not HAS_OPENPYXL:
        raise HTTPException(status_code=500, detail="openpyxl 未安装")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "角色导入"

    # 样式
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # 标题行
    for col, h in enumerate(_XLSX_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 示例数据（2 行，展示同一角色多时间节点）
    for r, row_data in enumerate(_XLSX_SAMPLE, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # 列宽
    col_widths = [14, 10, 8, 8, 10, 30, 10, 16, 30, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 冻结首行
    ws.freeze_panes = "A2"

    # Sheet2: 说明
    ws2 = wb.create_sheet("填写说明")
    notes = [
        ["字段", "说明"],
        ["小说", "必须与系统中已有小说名称一致"],
        ["姓名", "角色姓名，与年龄一起作为唯一标识"],
        ["性别", "男性 或 女性"],
        ["年龄", "整数，≥18"],
        ["身份", "如：将军、军师、士兵、读书人、文官、小吏、商人"],
        ["角色背景", "可选，角色的背景故事描述"],
        ["初始积分", "整数，默认0"],
        ["时间节点", "该角色所属的时间节点名称"],
        ["时间节点背景", "可选，该时间节点下的角色背景"],
        ["初始场景描述", "该时间节点下的开场场景文字"],
        ["", ""],
        ["注意", "同一角色如有多个时间节点，请在多行中重复填写（小说/姓名/性别/年龄/身份/背景/积分保持一致）"],
        ["导入规则", "姓名+年龄相同的视为同一角色，导入时会替换已有角色的全部数据"],
    ]
    for r, row_data in enumerate(notes, 1):
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            if r == 1:
                cell.font = Font(bold=True)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=character_import_template.xlsx"},
    )


@router.post("/character-import")
async def import_characters(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    _: bool = Depends(_verify_admin),
):
    """上传 xlsx 批量导入角色"""
    if not HAS_OPENPYXL:
        raise HTTPException(status_code=500, detail="openpyxl 未安装")

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="请上传 .xlsx 文件")

    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"无法读取 xlsx 文件: {e}")

    rows = list(ws.iter_rows(min_row=2, values_only=True))  # 跳过标题行
    if not rows:
        raise HTTPException(status_code=400, detail="文件中没有数据")

    # 按 (novel, name, age) 分组
    groups: dict = {}
    for row in rows:
        if not row[0] or not row[1]:
            continue  # 跳过空行
        novel = str(row[0]).strip()
        name = str(row[1]).strip()
        gender = str(row[2]).strip() if row[2] else None
        try:
            age = int(row[3]) if row[3] else 0
        except (ValueError, TypeError):
            age = 0
        rank = str(row[4]).strip() if row[4] else None
        background = str(row[5]).strip() if row[5] else None
        try:
            starting_points = int(row[6]) if row[6] else 0
        except (ValueError, TypeError):
            starting_points = 0
        timeline_name = str(row[7]).strip() if row[7] else None
        timeline_bg = str(row[8]).strip() if row[8] else None
        initial_scene = str(row[9]).strip() if row[9] else ""

        key = (novel, name, age)
        if key not in groups:
            groups[key] = {
                "novel": novel,
                "name": name,
                "gender": gender,
                "age": age,
                "rank": rank,
                "background": background,
                "starting_points": starting_points,
                "timelines": [],
            }
        if timeline_name:
            groups[key]["timelines"].append({
                "timeline": timeline_name,
                "background": timeline_bg,
                "initial_scene": initial_scene,
            })

    created = 0
    updated = 0
    errors = []

    for key, data in groups.items():
        novel, name, age = key
        try:
            # 查找是否已存在同名同年龄角色
            result = await db.execute(
                select(PresetCharacter).options(selectinload(PresetCharacter.timelines))
                .where(
                    PresetCharacter.novel == data["novel"],
                    PresetCharacter.name == data["name"],
                    PresetCharacter.age == data["age"],
                )
            )
            existing = result.scalar_one_or_none()

            if existing:
                # 更新已有角色
                existing.novel = data["novel"]
                existing.gender = data["gender"]
                existing.rank = data["rank"]
                existing.background = data["background"]
                existing.starting_points = data["starting_points"]
                # 替换时间线条目
                for old_tl in list(existing.timelines):
                    await db.delete(old_tl)
                for tl in data["timelines"]:
                    existing.timelines.append(CharacterTimeline(
                        timeline=tl["timeline"],
                        background=tl["background"],
                        initial_scene=tl["initial_scene"],
                    ))
                updated += 1
            else:
                # 创建新角色
                char = PresetCharacter(
                    novel=data["novel"], name=data["name"],
                    gender=data["gender"], age=data["age"],
                    rank=data["rank"], background=data["background"],
                    starting_points=data["starting_points"],
                )
                for tl in data["timelines"]:
                    char.timelines.append(CharacterTimeline(
                        timeline=tl["timeline"],
                        background=tl["background"],
                        initial_scene=tl["initial_scene"],
                    ))
                db.add(char)
                created += 1
        except Exception as e:
            errors.append(f"{name}({novel}): {e}")

    await db.commit()

    return {
        "message": f"导入完成：新建 {created} 个角色，更新 {updated} 个角色",
        "created": created,
        "updated": updated,
        "errors": errors,
    }
